
from django.contrib import admin
from pesquisa_satisfacao import models
import openpyxl
from django.http import HttpResponse
import re



@admin.register(models.Feedback)
class FeedbackAdmin(admin.ModelAdmin):
    list_display = 'id', 'email', 'rating', 'selected_options', 'indicacao', 'pedido', 'timestamp', 'data_pedido', 'loja_id'
    ordering = '-id',
    list_editable = 'data_pedido',
    # list_filter = 'created_date',
    search_fields = 'id', 'email', 'rating', 'selected_options', 'indicacao', 'pedido', 'timestamp'
    list_per_page = 100
    list_max_show_all = 200
    def export_to_excel(self, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="feedback.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        headers = ['ID', 'Rating', 'Selected Options', 'Indicação', 'pedido', 'timestamp', 'data_pedido', 'loja_id']
        for col_num, header_title in enumerate(headers, 1):
            column_letter = openpyxl.utils.get_column_letter(col_num)
            worksheet[f'{column_letter}1'] = header_title

       
        next_row = 2 

        for feedback in queryset:
            worksheet[f'A{next_row}'] = feedback.id
            worksheet[f'B{next_row}'] = feedback.rating

            selected_options = feedback.selected_options.split(',')  
            selected_options = [re.sub(r"[\[\]']",'', option) for option in selected_options] 

            first_mark_row = next_row 

            for index, option in enumerate(selected_options, start=next_row):
                worksheet.cell(row=index, column=3, value=option)  
                worksheet.cell(row=index, column=1, value=feedback.id)  
                worksheet.cell(row=index, column=2, value=feedback.rating)  
                worksheet.cell(row=index, column=4, value=feedback.indicacao) 
                worksheet.cell(row=index, column=5, value=feedback.pedido)  
                worksheet.cell(row=index, column=6, value=feedback.timestamp) 
                worksheet.cell(row=index, column=7, value=feedback.data_pedido) 
                worksheet.cell(row=index, column=8, value=feedback.loja_id) 

                next_row += 1

        workbook.save(response)
        return response

    export_to_excel.short_description = 'Exportar para Excel'
    actions = [export_to_excel]